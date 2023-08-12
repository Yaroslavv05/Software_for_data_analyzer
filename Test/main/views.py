from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import MyForm
from binance.client import Client
from datetime import datetime, timedelta
import openpyxl
from django.http import HttpResponse
from .models import MyData
import asyncio
import httpx
import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import FormDataSerializer
from .tasks import process_data_async


def main(request):
    return render(request, 'main.html')


def minute(symbol, open_price, bound, date, time_frame):
    client = Client()
    interval = Client.KLINE_INTERVAL_1MINUTE
    start_date = date
    start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S') - timedelta(hours=2)
    end_date_datetime = start_date_datetime + timedelta(hours=time_frame)

    klines = client.futures_historical_klines(symbol, interval, start_date_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                              end_date_datetime.strftime('%Y-%m-%d %H:%M:%S'))

    mass = []

    for kline in klines:
        time = datetime.fromtimestamp(kline[0] / 1000)
        formatted_time = time.strftime('%Y-%m-%d %H:%M:%S')
        mass.append({
            'time': formatted_time,
            'open': kline[1],
            'high': kline[2],
            'low': kline[3],
        })
    open_price = open_price
    bound = bound
    for i in mass:
        if (float(i['high']) - open_price) >= bound:
            return '1'
        elif open_price - float(i['low']) >= bound:
            return '0'


def get_binance_symbols():
    client = Client()
    exchange_info = client.get_exchange_info()
    symbols = [symbol['symbol'] for symbol in exchange_info['symbols']]
    return symbols


def index(request):
    form = MyForm(request.POST)
    if request.method == 'POST':
        if form.is_valid():
            symbol = form.cleaned_data['symbol']
            interval = form.cleaned_data['interval']
            bound = form.cleaned_data['bound']
            bound_unit = form.cleaned_data['bound_unit']
            start_data = form.cleaned_data['start_data']
            end_data = form.cleaned_data['end_data']
            data = {
                'symbol': symbol,
                'interval': interval,
                'bound': bound,
                'bound_unit': bound_unit,
                'start_data': start_data,
                'end_data': end_data
            }
            print(data)
            process_data_async.delay(data)
            return redirect('process')
    else:
        form = MyForm()
    return render(request, 'index.html', {'form': form})


def process(request):
    return render(request, 'proces.html')


# class FormSubmissionView(APIView):
#     def post(self, request, format=None):
#         serializer = FormDataSerializer(data=request.data)
#         if serializer.is_valid():
#             client = Client()
#             symbol = serializer.validated_data['symbol'].upper()
#             timeframe = float(serializer.validated_data['interval']) * 60
#             bound = float(serializer.validated_data['bound'])
#             bound_unit = serializer.validated_data['bound_unit']
#             inter = serializer.validated_data['interval']
#             print(symbol, timeframe, bound, bound_unit, inter)
#             interval_mapping = {
#                 0.0166666667: Client.KLINE_INTERVAL_1MINUTE,
#                 0.05: Client.KLINE_INTERVAL_3MINUTE,
#                 0.0833333333: Client.KLINE_INTERVAL_5MINUTE,
#                 0.25: Client.KLINE_INTERVAL_15MINUTE,
#             }
#             # if inter == '0.0166666667':
#             #     interval = Client.KLINE_INTERVAL_1MINUTE
#             # elif inter == '0.05':
#             #     interval = Client.KLINE_INTERVAL_3MINUTE
#             # elif inter == '0.0833333333':
#             #     interval = Client.KLINE_INTERVAL_5MINUTE
#             # elif inter == '0.25':
#             #     interval = Client.KLINE_INTERVAL_15MINUTE
#             # elif inter == '0.5':
#             #     interval = Client.KLINE_INTERVAL_30MINUTE
#             # elif inter == '1.0':
#             #     interval = Client.KLINE_INTERVAL_1HOUR
#             # elif inter == '2.0':
#             #     interval = Client.KLINE_INTERVAL_2HOUR
#             # elif inter == '4.0':
#             #     interval = Client.KLINE_INTERVAL_4HOUR
#             # elif inter == '6.0':
#             #     interval = Client.KLINE_INTERVAL_6HOUR
#             # elif inter == '8.0':
#             #     interval = Client.KLINE_INTERVAL_8HOUR
#             # elif inter == '12.0':
#             #     interval = Client.KLINE_INTERVAL_12HOUR
#             # elif inter == '24.0':
#             #     interval = Client.KLINE_INTERVAL_1DAY
#             # elif inter == '72.0':
#             #     interval = Client.KLINE_INTERVAL_3DAY
#             # elif inter == '168.0':
#             #     interval = Client.KLINE_INTERVAL_1WEEK
#             # elif inter == '720.0':
#             #     interval = Client.KLINE_INTERVAL_1MONTH
#             # else:
#             #     print('ggggg')
#             #     return HttpResponse('GG')
#
#             start_date_str = serializer.validated_data['start_data']
#             if 'T' in start_date_str:
#                 start_date = datetime.fromisoformat(start_date_str)
#             else:
#                 start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
#
#             end_date_str = serializer.validated_data['end_data']
#             if 'T' in end_date_str:
#                 end_date = datetime.fromisoformat(end_date_str)
#             else:
#                 end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
#                 end_date = end_date + timedelta(days=1)
#
#
#             difference = end_date - start_date
#             num_days = int(difference.days)
#             minutes = num_days * 24 * 60
#             hours = int(minutes / timeframe)
#
#             start_timestamp = int(start_date.timestamp()) * 1000
#             end_timestamp = int(end_date.timestamp()) * 1000
#
#             klines = client.futures_historical_klines(symbol, interval_mapping.get(inter), start_timestamp, end_timestamp)
#
#             mass = []
#
#             for kline in klines:
#                 time = datetime.fromtimestamp(kline[0] / 1000)
#                 formatted_time = time.strftime('%Y-%m-%d %H:%M:%S')
#                 mass.append({
#                     'time': formatted_time,
#                     'open': kline[1],
#                     'high': kline[2],
#                     'low': kline[3],
#                 })
#             output_data = []
#             total = 0
#             if bound_unit == '$':
#                 for i in range(hours):
#                     if i == 0:
#                         if (float(mass[i]['high']) - float(mass[i]['open'])) >= bound and float(
#                                 mass[i]['open']) - float(mass[i]['low']) >= bound:
#                             time = mass[i]['time']
#                             output = minute(symbol=symbol, open_price=float(mass[i]['open']), bound=bound, date=time,
#                                             time_frame=float(serializer.validated_data['interval']))
#                         elif (float(mass[i]['high']) - float(mass[i]['open'])) >= bound:
#                             time = mass[i]['time']
#                             output = '1'
#                         elif float(mass[i]['open']) - float(mass[i]['low']) >= bound:
#                             time = mass[i]['time']
#                             output = '0'
#                         else:
#                             output = '2'
#                             time = mass[i]['time']
#                         print(mass[total])
#                     else:
#                         total += int(timeframe)
#                         if (float(mass[i]['high']) - float(mass[i]['open'])) >= bound and float(
#                                 mass[i]['open']) - float(mass[i]['low']) >= bound:
#                             time = mass[i]['time']
#                             output = minute(symbol=symbol, open_price=float(mass[i]['open']), bound=bound, date=time,
#                                             time_frame=float(serializer.validated_data['interval']))
#                         elif (float(mass[i]['high']) - float(mass[i]['open'])) >= bound:
#                             time = mass[i]['time']
#                             output = '1'
#                         elif float(mass[i]['open']) - float(mass[i]['low']) >= bound:
#                             time = mass[i]['time']
#                             output = '0'
#                         else:
#                             output = '2'
#                             time = mass[i]['time']
#                         print(mass[i])
#                     output_data.append({'time': time, 'output': output})
#             elif bound_unit == '%':
#                 for i in range(hours):
#                     if i == 0:
#                         if (float(mass[i]['high']) - float(mass[i]['open'])) >= (
#                                 float(mass[total]['open']) / 100 * bound) and float(mass[i]['open']) - float(
#                             mass[i]['low']) >= (float(mass[total]['open']) / 100 * bound):
#                             time = mass[i]['time']
#                             output = minute(symbol=symbol, open_price=float(mass[i]['open']),
#                                             bound=(float(mass[i]['open']) / 100 * bound), date=time,
#                                             time_frame=float(serializer.validated_data['interval']))
#                         elif (float(mass[i]['high']) - float(mass[i]['open'])) >= (
#                                 float(mass[i]['open']) / 100 * bound):
#                             output = '1'
#                             time = mass[i]['time']
#                         elif float(mass[i]['open']) - float(mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
#                             output = '0'
#                             time = mass[i]['time']
#                         else:
#                             output = '2'
#                             time = mass[i]['time']
#                         print(mass[total])
#                     else:
#                         total += int(timeframe)
#                         if (float(mass[i]['high']) - float(mass[i]['open'])) >= (
#                                 float(mass[i]['open']) / 100 * bound) and float(mass[i]['open']) - float(
#                             mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
#                             time = mass[i]['time']
#                             output = minute(symbol=symbol, open_price=float(mass[i]['open']),
#                                             bound=(float(mass[i]['open']) / 100 * bound), date=time,
#                                             time_frame=float(serializer.validated_data['interval']))
#                         elif (float(mass[i]['high']) - float(mass[i]['open'])) >= (
#                                 float(mass[i]['open']) / 100 * bound):
#                             output = '1'
#                             time = mass[i]['time']
#                         elif float(mass[i]['open']) - float(mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
#                             output = '0'
#                             time = mass[i]['time']
#                         else:
#                             output = '2'
#                             time = mass[i]['time']
#                         print(mass[i])
#
#                     output_data.append({'time': time, 'output': output})
#             wb = openpyxl.Workbook()
#             ws = wb.active
#             daily_data = {}
#             day_count = 1
#             for item in output_data:
#                 day = item['time'][:10]
#                 if day in daily_data:
#                     daily_data[day].append(item)
#                 else:
#                     daily_data[day] = [item]
#
#             for day_data in daily_data.values():
#                 for col_index, item in enumerate(day_data, 1):
#                     ws.cell(row=day_count, column=1, value=datetime.strptime(item['time'], "%Y-%m-%d %H:%M:%S").date())
#                     ws.cell(row=day_count, column=col_index + 1, value=item['output'])
#                 day_count += 1
#             file_path = f"{symbol} data.xlsx"
#             wb.save(file_path)
#             return Response("Данные успешно обработаны.")
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def result(request):
    file_name = request.session.get('file_name', None)
    if file_name and os.path.exists(file_name):
        with open(file_name, 'rb') as file:
            response = HttpResponse(file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_name)}"'
            return response
    else:
        return HttpResponse("File not found.")
