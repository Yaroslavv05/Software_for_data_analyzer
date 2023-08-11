
from binance.client import Client
from datetime import datetime, timedelta
import openpyxl
from models import MyData


def minute(symbol, open_price, bound, date, time_frame):
    client = Client()
    interval = Client.KLINE_INTERVAL_1MINUTE
    start_date = date
    start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S') - timedelta(hours=2)
    end_date_datetime = start_date_datetime + timedelta(hours=time_frame)

    klines = client.futures_historical_klines(symbol, interval, start_date_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                              end_date_datetime.strftime('%Y-%m-%d %H:%M:%S'))

    mass = []

    for kline in klines:
        time = datetime.fromtimestamp(kline[0] / 1000)
        formatted_time = time.strftime('%Y-%m-%d %H:%M:%S')
        mass.append({
            'time': formatted_time,
            'open': kline[1],
            'high': kline[2],
            'low': kline[3],
        })
    open_price = open_price
    bound = bound
    for i in mass:
        if (float(i['high']) - open_price) >= bound:
            return '1'
        elif open_price - float(i['low']) >= bound:
            return '0'


latest_object = MyData.objects.latest('id')
client = Client()
symbol = latest_object.symbol.upper()
timeframe = float(latest_object.interval) * 60
bound = float(latest_object.bound)
bound_unit = latest_object.bound_unit
inter = latest_object.interval
print(inter)
if inter == '0.0166666667':
    interval = Client.KLINE_INTERVAL_1MINUTE
elif inter == '0.05':
    interval = Client.KLINE_INTERVAL_3MINUTE
elif inter == '0.0833333333':
    interval = Client.KLINE_INTERVAL_5MINUTE
elif inter == '0.25':
    interval = Client.KLINE_INTERVAL_15MINUTE
elif inter == '0.5':
    interval = Client.KLINE_INTERVAL_30MINUTE
elif inter == '1.0':
    interval = Client.KLINE_INTERVAL_1HOUR
elif inter == '2.0':
    interval = Client.KLINE_INTERVAL_2HOUR
elif inter == '4.0':
    interval = Client.KLINE_INTERVAL_4HOUR
elif inter == '6.0':
    interval = Client.KLINE_INTERVAL_6HOUR
elif inter == '8.0':
    interval = Client.KLINE_INTERVAL_8HOUR
elif inter == '12.0':
    interval = Client.KLINE_INTERVAL_12HOUR
elif inter == '24.0':
    interval = Client.KLINE_INTERVAL_1DAY
elif inter == '72.0':
    interval = Client.KLINE_INTERVAL_3DAY
elif inter == '168.0':
    interval = Client.KLINE_INTERVAL_1WEEK
elif inter == '720.0':
    interval = Client.KLINE_INTERVAL_1MONTH
else:
    print('gg')

start_date_str = latest_object.start_data
if 'T' in start_date_str:
    start_date = datetime.fromisoformat(start_date_str)
else:
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')

end_date_str = latest_object.end_data
if 'T' in end_date_str:
    end_date = datetime.fromisoformat(end_date_str)
else:
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    end_date = end_date + timedelta(days=1)

difference = end_date - start_date
num_days = int(difference.days)
minutes = num_days * 24 * 60
hours = int(minutes / timeframe)

start_timestamp = int(start_date.timestamp()) * 1000
end_timestamp = int(end_date.timestamp()) * 1000

klines = client.futures_historical_klines(symbol, interval, start_timestamp, end_timestamp)

mass = []

for kline in klines:
    time = datetime.fromtimestamp(kline[0] / 1000)
    formatted_time = time.strftime('%Y-%m-%d %H:%M:%S')
    mass.append({
        'time': formatted_time,
        'open': kline[1],
        'high': kline[2],
        'low': kline[3],
    })
output_data = []
total = 0
if bound_unit == '$':
    for i in range(hours):
        if i == 0:
            if (float(mass[i]['high']) - float(mass[i]['open'])) >= bound and float(
                    mass[i]['open']) - float(mass[i]['low']) >= bound:
                time = mass[i]['time']
                output = minute(symbol=symbol, open_price=float(mass[i]['open']), bound=bound, date=time,
                                time_frame=float(latest_object.interval))
            elif (float(mass[i]['high']) - float(mass[i]['open'])) >= bound:
                time = mass[i]['time']
                output = '1'
            elif float(mass[i]['open']) - float(mass[i]['low']) >= bound:
                time = mass[i]['time']
                output = '0'
            else:
                output = '2'
                time = mass[i]['time']
            print(mass[total])
        else:
            total += int(timeframe)
            if (float(mass[i]['high']) - float(mass[i]['open'])) >= bound and float(
                    mass[i]['open']) - float(mass[i]['low']) >= bound:
                time = mass[i]['time']
                output = minute(symbol=symbol, open_price=float(mass[i]['open']), bound=bound, date=time,
                                time_frame=float(latest_object.interval))
            elif (float(mass[i]['high']) - float(mass[i]['open'])) >= bound:
                time = mass[i]['time']
                output = '1'
            elif float(mass[i]['open']) - float(mass[i]['low']) >= bound:
                time = mass[i]['time']
                output = '0'
            else:
                output = '2'
                time = mass[i]['time']
            print(mass[i])
        output_data.append({'time': time, 'output': output})
elif bound_unit == '%':
    for i in range(hours):
        if i == 0:
            if (float(mass[i]['high']) - float(mass[i]['open'])) >= (
                    float(mass[total]['open']) / 100 * bound) and float(mass[i]['open']) - float(
                mass[i]['low']) >= (float(mass[total]['open']) / 100 * bound):
                time = mass[i]['time']
                output = minute(symbol=symbol, open_price=float(mass[i]['open']),
                                bound=(float(mass[i]['open']) / 100 * bound), date=time,
                                time_frame=float(latest_object.interval))
            elif (float(mass[i]['high']) - float(mass[i]['open'])) >= (
                    float(mass[i]['open']) / 100 * bound):
                output = '1'
                time = mass[i]['time']
            elif float(mass[i]['open']) - float(mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
                output = '0'
                time = mass[i]['time']
            else:
                output = '2'
                time = mass[i]['time']
            print(mass[total])
        else:
            total += int(timeframe)
            if (float(mass[i]['high']) - float(mass[i]['open'])) >= (
                    float(mass[i]['open']) / 100 * bound) and float(mass[i]['open']) - float(
                mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
                time = mass[i]['time']
                output = minute(symbol=symbol, open_price=float(mass[i]['open']),
                                bound=(float(mass[i]['open']) / 100 * bound), date=time,
                                time_frame=float(latest_object.interval))
            elif (float(mass[i]['high']) - float(mass[i]['open'])) >= (
                    float(mass[i]['open']) / 100 * bound):
                output = '1'
                time = mass[i]['time']
            elif float(mass[i]['open']) - float(mass[i]['low']) >= (float(mass[i]['open']) / 100 * bound):
                output = '0'
                time = mass[i]['time']
            else:
                output = '2'
                time = mass[i]['time']
            print(mass[i])

        output_data.append({'time': time, 'output': output})
wb = openpyxl.Workbook()
ws = wb.active
daily_data = {}
day_count = 1
for item in output_data:
    day = item['time'][:10]
    if day in daily_data:
        daily_data[day].append(item)
    else:
        daily_data[day] = [item]

for day_data in daily_data.values():
    for col_index, item in enumerate(day_data, 1):
        ws.cell(row=day_count, column=1, value=datetime.strptime(item['time'], "%Y-%m-%d %H:%M:%S").date())
        ws.cell(row=day_count, column=col_index + 1, value=item['output'])
    day_count += 1
file_path = f"{symbol} data.xlsx"
wb.save(file_path)
