import requests
from datetime import datetime, timezone, timedelta
import pytz
import openpyxl
import os
from ..models import DateLog

class FormatingDataService:
    def __init__(self, symbol, bound_up, bound_unit_up, bound_low, bound_unit_low, start_date, end_date, min_interval, api_key, asset_type):
        self.symbol = symbol
        self.bound_up = float(bound_up)
        self.bound_unit_up = bound_unit_up
        self.bound_low = float(bound_low)
        self.bound_unit_low = bound_unit_low
        self.start_date = start_date
        self.end_date = end_date
        self.min_interval = min_interval
        self.api_key = api_key
        self.asset_type = asset_type


    def second(self, symbol, open_price, date, bound_up, bound_low):
            start_date = date
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
            if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
                start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
            else:
                start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
            ny_timezone = pytz.timezone('America/New_York')
            start_date_datetime = ny_timezone.localize(start_date_datetime)
            end_date_datetime = start_date_datetime + timedelta(hours=1)
            start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
            end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000

            if self.asset_type == 'currency':
                response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/C:{symbol}/range/1/second/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}")
            else:
                response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/second/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}")

            print(response.json())
            d = response.json()['results']
            mass = []
            for i in d:
                dt = datetime.fromtimestamp(i['t'] / 1000)
                mass.append({
                    'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'open': i['o'],
                    'close': i['c'],
                    'high': i['h'],
                    'low': i['l']
                })

            for i in mass:
                if float(i['high']) - open_price >= bound_up:
                    return '1'
                elif open_price - float(i['low']) >= bound_low:
                    return '0'

    def minte(self, date, symbol, open_price, bound_up, bound_low):
        start_date = date
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=1)
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000

        if self.asset_type == 'currency':
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/C:{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}")
        else:
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}")

        print(response.json())
        d = response.json()['results']
        mass = []
        for i in d:
            dt = datetime.fromtimestamp(i['t'] / 1000)
            mass.append({
                'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'open': i['o'],
                'close': i['c'],
                'high': i['h'],
                'low': i['l']
            })

        for i in mass:
            if float(i['high']) - open_price >= bound_up:
                return '1'
            elif open_price - float(i['low']) >= bound_low:
                return '0'


    def is_valid_time(self, time):
            return (time.hour == 9 and time.minute >= 30) or (9 < time.hour < 15) or (time.hour == 15 and time.minute <= 30)


    def convert_unix_to_datetime(self, unix_timestamp):
        dt_object = datetime.fromtimestamp(unix_timestamp / 1000.0, tz=timezone.utc)
        dt_object = dt_object.astimezone(timezone(timedelta(hours=-5)))
        return dt_object.strftime('%Y-%m-%d %H:%M:%S')


    def create_4h_candles(self, data):
        print(data)
        candles_4h = []

        current_candle = None

        for entry in data:
            entry_time = datetime.strptime(entry['t'], '%Y-%m-%d %H:%M:%S')

            if current_candle is None:
                current_candle = {
                    'o': entry['o'],
                    'h': entry['h'],
                    'l': entry['l'],
                    'c': entry['c'],
                    'v': entry['v'],
                    't': entry_time.replace(minute=30, second=0),
                    'n': entry['n']
                }
            else:
                if entry_time >= current_candle['t'] + timedelta(hours=4):
                    current_candle['t'] = current_candle['t'].strftime('%Y-%m-%d %H:%M:%S')
                    candles_4h.append(current_candle)
                    current_candle = {
                        'o': entry['o'],
                        'h': entry['h'],
                        'l': entry['l'],
                        'c': entry['c'],
                        'v': entry['v'],
                        't': entry_time.replace(minute=30, second=0),
                        'n': entry['n']
                    }
                else:
                    current_candle['h'] = max(current_candle['h'], entry['h'])
                    current_candle['l'] = min(current_candle['l'], entry['l'])
                    current_candle['c'] = entry['c']
                    current_candle['v'] += entry['v']
                    current_candle['n'] += entry['n']

        if current_candle:
            current_candle['t'] = current_candle['t'].strftime('%Y-%m-%d %H:%M:%S')
            candles_4h.append(current_candle)

        return candles_4h


    def split_into_3_month_intervals(self, start_date, end_date):
        intervals = []
        current_date = start_date
        while current_date < end_date:
            next_date = current_date + timedelta(days=90)
            if next_date > end_date:
                next_date = end_date
            intervals.append((current_date, next_date))
            current_date = next_date
        return intervals

    def output(self):
        intervals = self.split_into_3_month_intervals(datetime.strptime(self.start_date, '%Y-%m-%d').date(), datetime.strptime(self.end_date, '%Y-%m-%d').date())

        results = []
        for i, interval in enumerate(intervals, start=1):
            if self.asset_type == 'currency':
                url = f'https://api.polygon.io/v2/aggs/ticker/C:{self.symbol}/range/30/minute/{interval[0].strftime("%Y-%m-%d")}/{interval[1].strftime("%Y-%m-%d")}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}'
            else:
                url = f'https://api.polygon.io/v2/aggs/ticker/{self.symbol}/range/30/minute/{interval[0].strftime("%Y-%m-%d")}/{interval[1].strftime("%Y-%m-%d")}?adjusted=true&sort=asc&limit=50000&apiKey={self.api_key}'

            response = requests.get(url)
            results.append(response.json()['results'])

        data = []
        for result in results:
            for j in result:
                j['t'] = self.convert_unix_to_datetime(j['t'])
                time = datetime.strptime(j['t'], '%Y-%m-%d %H:%M:%S')
                if self.is_valid_time(time):
                    data.append(j)

        candles_4h = self.create_4h_candles(data)

        self.output_data = []

        for candle in candles_4h:
            parsed_date = datetime.strptime(candle['t'], "%Y-%m-%d %H:%M:%S")
            DateLog.objects.create(date=parsed_date.date(), task_id='1')
            if self.bound_unit_up == '$':
                bound_value_up = self.bound_up
            elif self.bound_unit_up == '%':
                bound_value_up = (float(candle['o']) / 100) * self.bound_up
            
            if self.bound_unit_low == '$':
                bound_value_low = self.bound_low
            elif self.bound_unit_low == '%':
                bound_value_low = (float(candle['o']) / 100) * self.bound_low

            if float(candle['h']) - float(candle['o']) >= bound_value_up and float(candle['o']) - float(candle['l']) >= bound_value_low:
                if self.min_interval == '60':
                    output = self.minte(symbol=self.symbol, date=candle['t'], open_price=float(candle['o']), bound_up=bound_value_up, bound_low=bound_value_low)
                elif self.min_interval == '1':
                    output = self.second(symbol=self.symbol, date=candle['t'], open_price=float(candle['o']), bound_up=bound_value_up, bound_low=bound_value_low)
            elif float(candle['h']) - float(candle['o']) >= bound_value_up:
                output = '1'
            elif float(candle['o']) - float(candle['l']) >= bound_value_low:
                output = '0'
            else:
                output = '2'

            self.output_data.append({
                'time': candle['t'],
                'output': output,
                'open': candle['o'],
                'close': candle['c'],
                'high': candle['h'],
                'low': candle['l'],
                'trade': candle['n'],
                'volume': candle['v']
            })
            import time
            time.sleep(0.01)
            try:
                date_log = DateLog.objects.get(task_id='1')
                date_log.delete()
            except:
                print('Ничего не найденно по такому ID')

        return self.output_data
    
    
    def save_output_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Trade', 'Volume']
        for col_index, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_index, value=header)

        output_data = self.output()
        for item in output_data:
            row_data = [item['time'], item['output'], item['open'], item['close'], item['high'], item['low'], item['trade'], item['volume']]
            ws.append(row_data)
        file = f'{self.symbol}_4h_{self.bound_up}{self.bound_unit_up}_{self.bound_low}{self.bound_unit_low}_{self.start_date}_{self.end_date}(Polygon).xlsx'
        file = file.replace(':', '_').replace('?', '_').replace(' ', '_')
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'..\\..\\{file}')
        with open(file_path, 'wb') as file:
            wb.save(file)
        
        return file_path, output_data


class DataProcessor:
    def __init__(self, input_file_path, output_file_path, bound_up, bound_low, symbol, bound_unit_up, bound_unit_low, min_interval):
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.bound_up = bound_up
        self.bound_unit_up = bound_unit_up
        self.bound_low = bound_low
        self.bound_unit_low = bound_unit_low
        self.symbol = symbol
        self.min_interval = min_interval
        self.hourly_intervals = []

    def load_data(self):
        workbook = openpyxl.load_workbook(self.input_file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data_list = []

        for row in sheet.iter_rows(values_only=True, min_row=2):
            row_data = dict(zip(headers, row))
            time = datetime.strptime(row_data['Date'], '%Y-%m-%d %H:%M:%S')
            if self.is_valid_time(time):
                data_list.append(row_data)

        data_list.sort(key=lambda x: (x['Date'][-5:] == ':30', x['Date']))
        workbook.close()
        self.process_intervals(data_list)

    def is_valid_time(self, time):
        return (time.hour == 9 and time.minute >= 30) or (9 < time.hour < 15) or (time.hour == 15 and time.minute <= 30)

    def process_intervals(self, data_list):
        current_interval = None

        for candle in data_list:
            candle_date = datetime.strptime(candle['Date'], '%Y-%m-%d %H:%M:%S')

            if current_interval is None:
                current_interval = {
                    'time': candle_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'open': candle['Open'],
                    'close': candle['Close'],
                    'high': candle['High'],
                    'low': candle['Low'],
                    'trade': candle['Trade'],
                    'volume': candle['Volume']
                }
            else:
                current_interval_date = datetime.strptime(current_interval['time'], '%Y-%m-%d %H:%M:%S')
                time_difference = (candle_date - current_interval_date).total_seconds()
                if time_difference < 3600:
                    current_interval['close'] = candle['Close']
                    current_interval['high'] = max(current_interval['high'], candle['High'])
                    current_interval['low'] = min(current_interval['low'], candle['Low'])
                    current_interval['trade'] += candle['Trade']
                    current_interval['volume'] += candle['Volume']
                else:
                    self.hourly_intervals.append(current_interval)
                    current_interval = {
                        'time': candle_date.strftime('%Y-%m-%d %H:%M:%S'),
                        'open': candle['Open'],
                        'close': candle['Close'],
                        'high': candle['High'],
                        'low': candle['Low'],
                        'trade': candle['Trade'],
                        'volume': candle['Volume']
                    }

        if current_interval:
            self.hourly_intervals.append(current_interval)
    
    def second(self, symbol, open_price, date, bound_up, bound_low):
        start_date = date
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=1)
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000

        if self.asset_type == 'currency':
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/C:{symbol}/range/1/second/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")
        else:
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/second/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")

        print(response.json())
        d = response.json()['results']
        mass = []
        for i in d:
            dt = datetime.fromtimestamp(i['t'] / 1000)
            mass.append({
                'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'open': i['o'],
                'close': i['c'],
                'high': i['h'],
                'low': i['l']
            })

        for i in mass:
            if float(i['high']) - open_price >= bound_up:
                return '1'
            elif open_price - float(i['low']) >= bound_low:
                return '0'

    def minte(self, date, symbol, open_price, bound_up, bound_low):
        start_date = date
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=1)
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000

        if self.asset_type == 'currency':
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/C:{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")
        else:
            response = requests.get(f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")

        print(response.json())
        d = response.json()['results']
        mass = []
        for i in d:
            dt = datetime.fromtimestamp(i['t'] / 1000)
            mass.append({
                'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'open': i['o'],
                'close': i['c'],
                'high': i['h'],
                'low': i['l']
            })

        for i in mass:
            if float(i['high']) - open_price >= bound_up:
                return '1'
            elif open_price - float(i['low']) >= bound_low:
                return '0'

    def generate_output(self):
        output_data = []

        for i in self.hourly_intervals:
            if self.bound_unit_up == '$':
                bound_value_up = self.bound_up
            elif self.bound_unit_up == '%':
                bound_value_up = (float(i['open']) / 100) * self.bound_up
            
            if self.bound_unit_low == '$':
                bound_value_low = self.bound_low
            elif self.bound_unit_low == '%':
                bound_value_low = (float(i['open']) / 100) * self.bound_low

            if float(i['high']) - float(i['open']) >= bound_value_up and float(i['open']) - float(i['low']) >= bound_value_low:
                if self.min_interval == '60':
                    output = self.minte(symbol=self.symbol, date=i['time'], open_price=float(i['open']), bound_up=bound_value_up, bound_low=bound_value_low)
                elif self.min_interval == '1':
                    output = self.second(symbol=self.symbol, date=i['time'], open_price=float(i['open']), bound_up=bound_value_up, bound_low=bound_value_low)
            elif float(i['high']) - float(i['open']) >= bound_value_up:
                output = '1'
            elif float(i['open']) - float(i['low']) >= bound_value_low:
                output = '0'
            else:
                output = '2'

            output_data.append({
                'time': i['time'],
                'output': output,
                'open': i['open'],
                'close': i['close'],
                'high': i['high'],
                'low': i['low'],
                'trade': i['trade'],
                'volume': i['volume']
            })

        return output_data

    def save_output_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Trade', 'Volume']
        for col_index, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_index, value=header)

        output_data = self.generate_output()
        for item in output_data:
            row_data = [item['time'], item['output'], item['open'], item['close'], item['high'], item['low'], item['trade'], item['volume']]
            ws.append(row_data)

        with open(self.output_file_path, 'wb') as file:
            wb.save(file)