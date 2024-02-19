import io
from django.contrib.auth.models import User
from celery import shared_task
import yfinance as yf
import pandas as pd
from binance.client import Client
from datetime import datetime, timedelta, timezone
import openpyxl
import time
import requests
import os
import re
from .models import DataEntry
from plyer import notification
import pytz
import csv
from .models import *
from django.utils import timezone
from .services.formating_data_services import FormatingDataService, DataProcessor
from .services.formating_data_new_services import FormatingDataServiceNew


def minute(symbol, open_price, bound_up, bound_low, date, time_frame):
    client = Client()
    interval = Client.KLINE_INTERVAL_1MINUTE
    start_date = date
    start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    end_date_datetime = start_date_datetime + timedelta(hours=time_frame)

    klines = client.futures_historical_klines(symbol, interval, start_date_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                              end_date_datetime.strftime('%Y-%m-%d %H:%M:%S'))

    mass = []

    for kline in klines:
        time = datetime.fromtimestamp(kline[0] / 1000)
        formatted_time = time.strftime('%Y-%m-%d %H:%M:%S')
        mass.append({
            'time': formatted_time,
            'open': kline[1],
            'high': kline[2],
            'low': kline[3],
        })
    open_price = open_price
    for i in mass:
        if (float(i['high']) - open_price) >= bound_up:
            return '1'
        elif open_price - float(i['low']) >= bound_low:
            return '0'


@shared_task
def process_data_async(data):
    client = Client()
    symbol = data['symbol'].upper()
    timeframe = float(data['interval']) * 60
    bound_up = float(data['bound_up'])
    bound_unit_up = data['bound_unit_up']
    bound_low = float(data['bound_low'])
    bound_unit_low = data['bound_unit_low']
    inter = data['interval']
    
    interval_mapping = {
        0.0166666667: Client.KLINE_INTERVAL_1MINUTE,
        0.05: Client.KLINE_INTERVAL_3MINUTE,
        0.0833333333: Client.KLINE_INTERVAL_5MINUTE,
        0.25: Client.KLINE_INTERVAL_15MINUTE,
        0.5: Client.KLINE_INTERVAL_30MINUTE,
        1.0: Client.KLINE_INTERVAL_1HOUR,
        2.0: Client.KLINE_INTERVAL_2HOUR,
        4.0: Client.KLINE_INTERVAL_4HOUR,
        6.0: Client.KLINE_INTERVAL_6HOUR,
        8.0: Client.KLINE_INTERVAL_8HOUR,
        12.0: Client.KLINE_INTERVAL_12HOUR,
        24.0: Client.KLINE_INTERVAL_1DAY,
        72.0: Client.KLINE_INTERVAL_3DAY,
        168.0: Client.KLINE_INTERVAL_1WEEK,
        720.0: Client.KLINE_INTERVAL_1MONTH
    }

    start_date_str = data['start_data']
    start_date = datetime.fromisoformat(start_date_str) if 'T' in start_date_str else datetime.strptime(start_date_str, '%Y-%m-%d')

    end_date_str = data['end_data']
    end_date = datetime.fromisoformat(end_date_str) if 'T' in end_date_str else datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) 

    difference = end_date - start_date
    minutes = difference.days * 24 * 60
    hours = int(minutes / timeframe)
    
    start_timestamp = int(start_date.timestamp()) * 1000
    end_timestamp = int(end_date.timestamp()) * 1000
    print(start_timestamp)
    print(end_timestamp)

    klines = client.futures_historical_klines(symbol, interval_mapping.get(float(inter), None), start_timestamp, end_timestamp)

    mass = []
    for kline in klines:
        time = datetime.fromtimestamp(kline[0] / 1000).strftime('%Y-%m-%d %H:%M:%S')
        mass.append({
            'time': time,
            'open': kline[1],
            'high': kline[2],
            'low': kline[3],
            'close': kline[4],
            'volume': kline[5]
        })
        
    print(mass)
    output_data = []
    for i in mass:
        time = i['time']
        high = float(i['high'])
        ope = float(i['open'])
        low = float(i['low'])
        volume = i['volume']
        parsed_date = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
        aware_date = timezone.make_aware(parsed_date)
        DateLog.objects.create(date=aware_date.date(), task_id=process_data_async.request.id)
        
        if bound_unit_up == '$':
            bound_check_up = bound_up
        elif bound_unit_up == '%':
            bound_check_up = float(i['open']) / 100 * bound_up
        
        if bound_unit_low == '$':
            bound_check_low = bound_low
        elif bound_unit_low == '%':
            bound_check_low = float(i['open']) / 100 * bound_low

        if float(i['high']) - float(i['open']) >= bound_check_up and float(i['open']) - float(i['low']) >= bound_check_low:
            output = minute(symbol=symbol, open_price=ope, bound_up=bound_check_up, bound_low=bound_check_low,  date=time, time_frame=float(data['interval']))
        elif float(i['high']) - float(i['open']) >= bound_check_up:
            output = '1'
        elif float(i['open']) - float(i['low']) >= bound_check_low:
            output = '0'
        else:
            output = '2'

        close = float(i['close'])
        output_data.append({'time': time, 'output': output, 'open': str(ope), 'close': str(close), 'high': str(high), 'low': str(low), 'volume': str(volume)})

        import time
        time.sleep(0.01)
        try:
            date_log = DateLog.objects.get(task_id=process_data_async.request.id)
            date_log.delete()
        except:
            print('Nothing found for this ID')
                    
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Volume']
    for col_index, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_index, value=header)
    for item in output_data:
        row_data = [item['time'], item['output'], item['open'].replace(',', '.'), item['close'].replace(',', '.'), item['high'].replace(',', '.'), item['low'].replace(',', '.'),
                    item['volume'].replace(',', '.')]
        ws.append(row_data)
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    file_path = f'{symbol}_{timeframe}_{bound_up}{bound_unit_up}_{bound_low}_{bound_unit_low}_{start_date}_{end_date}(Binance).xlsx'
    file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
    with open(file_path, 'wb') as file:
        file.write(output_buffer.read())
    task = Task.objects.get(user=data['us'], is_running=True)
    task.is_running = False
    task.save()
    return file_path, output_data


def controversial(symbol, timeframe, open_price, date, bound_up, bound_low):
    interval_mapping = {
        '1min': 0.0166666667,
        '5min': 0.05,
        '15min': 0.0833333333,
        '30min': 0.25,
        '45min': 0.375,
        '1h': 1.0,
        '2h': 2.0,
        '4h': 4.0,
        '1day': 24.0,
        '1week': 168.0,
        '1month': 720.0
    }
    start_date = date
    if len(date) == 10:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    end_date_datetime = start_date_datetime + timedelta(hours=float(interval_mapping[timeframe]))
    print(symbol, timeframe, bound_low, bound_up, start_date, end_date_datetime)
    response = requests.get(
        f"https://api.twelvedata.com/time_series?apikey=7e1f42d9a4f743749ffa9e77958e06a4&interval=1min&symbol={symbol}&timezone=exchange&start_date={start_date}&end_date={end_date_datetime}")
    try:
        d = response.json()['values']
    except:
        return '3'

    for j in d[::-1]:
        if float(j['high']) - open_price >= bound_up:
            return '1'
        elif open_price - float(j['low']) >= bound_low:
            return '0'


@shared_task
def shared_async_task(data):
    symbol = data['symbol'].upper()
    timeframe = data['interval']
    bound_up = float(data['bound_up'])
    bound_unit_up = data['bound_unit_up']
    bound_low = float(data['bound_low'])
    bound_unit_low = data['bound_unit_low']
    start_date_input = data['start_data']
    end_date_input = data['end_data']
    user_id = data['us']

    if len(start_date_input) == 10:
        start_date_input += ' 00:00:00'

    if len(end_date_input) == 10:
        end_date_input += ' 00:00:00'

    start_date = datetime.strptime(start_date_input, '%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(end_date_input, '%Y-%m-%d %H:%M:%S') + timedelta(days=1)

    start_date_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
    end_date_str = end_date.strftime('%Y-%m-%d %H:%M:%S')
    print(symbol, timeframe, bound_up, bound_unit_up, bound_low, bound_unit_low, start_date, end_date)
    response = requests.get(
        f"https://api.twelvedata.com/time_series?apikey=7e1f42d9a4f743749ffa9e77958e06a4&interval={timeframe}&symbol={symbol}&timezone=exchange&start_date={start_date_str}&end_date={end_date_str}")

    data = response.json()['values']
    import time
    output_data = []
    for i in data:
        if timeframe == '1day' or timeframe == '1week' or timeframe == '1month':
            parsed_date = datetime.strptime(i['datetime'], "%Y-%m-%d")
        else:
            parsed_date = datetime.strptime(i['datetime'], "%Y-%m-%d %H:%M:%S")
        aware_date = timezone.make_aware(parsed_date)
        DateLog.objects.create(date=aware_date.date(), task_id=shared_async_task.request.id)

        times = i['datetime']
        ope = i['open']
        close = i['close']
        high = i['high']
        low = i['low']
        volume = i['volume']

        if bound_unit_up == '$':
            bound_check_up = bound_up
        elif bound_unit_up == '%':
            bound_check_up = float(i['open']) / 100 * bound_up
        
        if bound_unit_low == '$':
            bound_check_low = bound_low
        elif bound_unit_low == '%':
            bound_check_low = float(i['open']) / 100 * bound_low

        if float(i['high']) - float(i['open']) >= bound_check_up and float(i['open']) - float(i['low']) >= bound_check_low:
            output = controversial(symbol=symbol, timeframe=timeframe, open_price=float(i['open']), date=i['datetime'], bound_up=bound_check_up, bound_low=bound_check_low)
            time.sleep(10)
        elif float(i['high']) - float(i['open']) >= bound_check_up:
            output = '1'
        elif float(i['open']) - float(i['low']) >= bound_check_low:
            output = '0'
        else:
            output = '2'

        output_data.append({'time': times, 'output': output, 'open': ope, 'close': close, 'high': high, 'low': low, 'volume': volume})
        time.sleep(0.01)
        try:
            date_log = DateLog.objects.get(task_id=shared_async_task.request.id)
            date_log.delete()
        except:
            print('Ничего не найденно по такому ID')

    print(output_data)
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Volume']
    for col_index, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_index, value=header)

    for item in output_data[::-1]:
        row_data = [item['time'], item['output'], item['open'], item['close'], item['high'], item['low'],
                    item['volume']]
        ws.append(row_data)
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    file_path = f'{symbol}_{timeframe}_{bound_up}{bound_unit_up}_{bound_low}{bound_unit_low}_{start_date}_{end_date}(Twelvedata).xlsx'
    file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
    with open(file_path, 'wb') as file:
        file.write(output_buffer.read())
    task = Task.objects.get(user=user_id, is_running=True)
    task.is_running = False
    task.save()
    return file_path, output_data[::-1]


def second_shares_polygon(symbol, timeframe, open_price, date, bound_up, bound_low):
    interval_mapping = {
        '1 minute': 0.0166666667,
        '5 minute': 0.0833333333,
        '15 minute': 0.25,
        '30 minute': 0.5,
        '45 minute': 0.75,
        '1 hour': 1.0,
        '2 hour': 2.0,
        '3 hour': 3.0,
        '4 hour': 4.0,
        '5 hour': 5.0,
        '6 hour': 6.0,
        '7 hour': 7.0,
        '8 hour': 8.0,
        '9 hour': 9.0,
        '10 hour': 10.0,
        '11 hour': 11.0,
        '12 hour': 12.0,
        '1 day': 24.0,
        '1 week': 168.0,
        '1 month': 720.0,
        '1 year': 8760
    }
    start_date = date
    start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
        start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
    else:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    ny_timezone = pytz.timezone('America/New_York')
    start_date_datetime = ny_timezone.localize(start_date_datetime)
    end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
    start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
    end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000

    print('1 секунда')
    response = requests.get(
        f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/second/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")
    print(response.json())
    d = response.json()['results']
    mass = []
    for i in d:
        dt = datetime.fromtimestamp(i['t'] / 1000)
        mass.append({
            'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
            'open': i['o'],
            'close': i['c'],
            'high': i['h'],
            'low': i['l']
        })

    for i in mass:
        if float(i['high']) - open_price >= bound_up:
            return '1'
        elif open_price - float(i['low']) >= bound_low:
            return '0'


def minute_shares_polygon(symbol, timeframe, open_price, date, bound_up, bound_low):
    interval_mapping = {
        '1 minute': 0.0166666667,
        '5 minute': 0.0833333333,
        '15 minute': 0.25,
        '30 minute': 0.5,
        '45 minute': 0.75,
        '1 hour': 1.0,
        '2 hour': 2.0,
        '3 hour': 3.0,
        '4 hour': 4.0,
        '5 hour': 5.0,
        '6 hour': 6.0,
        '7 hour': 7.0,
        '8 hour': 8.0,
        '9 hour': 9.0,
        '10 hour': 10.0,
        '11 hour': 11.0,
        '12 hour': 12.0,
        '1 day': 24.0,
        '1 week': 168.0,
        '1 month': 720.0,
        '1 year': 8760
    }
    start_date = date
    start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
        start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
    else:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    ny_timezone = pytz.timezone('America/New_York')
    start_date_datetime = ny_timezone.localize(start_date_datetime)
    end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
    start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
    end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000
    
    print('1 минута')
    response = requests.get(
        f"https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz")
    print(response.json())
    d = response.json()['results']
    mass = []
    for i in d:
        dt = datetime.fromtimestamp(i['t'] / 1000)
        mass.append({
            'time': dt.strftime('%Y-%m-%d %H:%M:%S'),
            'open': i['o'],
            'close': i['c'],
            'high': i['h'],
            'low': i['l']
        })

    for i in mass:
        if float(i['high']) - open_price >= bound_up:
            return '1'
        elif open_price - float(i['low']) >= bound_low:
            return '0'


def split_into_3_month_intervals(start_date, end_date):
    intervals = []
    current_date = start_date
    while current_date < end_date:
        next_date = current_date + timedelta(days=90)
        if next_date > end_date:
            next_date = end_date
        intervals.append((current_date, next_date))
        current_date = next_date
    return intervals


@shared_task
def shares_polygon_async_task(data):
    symbol = data['symbol'].upper()
    timeframe = data['interval']
    bound_up = float(data['bound_up'])
    bound_unit_up = data['bound_unit_up']
    bound_low = float(data['bound_low'])
    bound_unit_low = data['bound_unit_low']
    start_date = data['start_data']
    end_date = data['end_data']
    api = data['api']
    print(start_date)
    print(timeframe)
    if timeframe == '4 hour':
        formating = FormatingDataService(symbol=symbol, bound_up=bound_up, bound_unit_up=bound_unit_up, bound_low=bound_low, bound_unit_low=bound_unit_low, start_date=start_date, end_date=end_date, min_interval=data['min_interval'], api_key=api)
        file_path, output_data = formating.save_output_to_excel()
        task = Task.objects.get(user=data['us'], is_running=True)
        task.is_running = False
        task.save()
        return file_path, output_data
    else:
        intervals = split_into_3_month_intervals(datetime.strptime(start_date, '%Y-%m-%d').date(), datetime.strptime(end_date, '%Y-%m-%d').date())

        res = []
        for i, interval in enumerate(intervals, start=1):
            if timeframe == '1 hour':
                url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/30/minute/{interval[0].strftime("%Y-%m-%d")}/{interval[1].strftime("%Y-%m-%d")}?adjusted=true&sort=asc&limit=50000&apiKey={api}'
            else:
                interval_parts = timeframe.split()
                url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/{interval_parts[0]}/{interval_parts[1]}/{interval[0].strftime("%Y-%m-%d")}/{interval[1].strftime("%Y-%m-%d")}?adjusted=true&sort=asc&limit=50000&apiKey={api}'

            response = requests.get(url)
            res.append(response.json()['results'])

        mass = []
        for subarray in res:
            for i in subarray:
                unix_timestamp_seconds = i['t'] / 1000
                unix_datetime = datetime.fromtimestamp(unix_timestamp_seconds, pytz.utc)
                ny_timezone = pytz.timezone('America/New_York')
                ny_datetime = unix_datetime.astimezone(ny_timezone)
                mass.append({
                    'time': ny_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                    'open': i['o'],
                    'close': i['c'],
                    'high': i['h'],
                    'low': i['l'],
                    'trade': i['n'],
                    'volume': i['v']
                })

        output_data = []
        for i in mass:
            parsed_date = datetime.strptime(i['time'], "%Y-%m-%d %H:%M:%S")
            aware_date = timezone.make_aware(parsed_date)
            DateLog.objects.create(date=aware_date.date(), task_id=shares_polygon_async_task.request.id)
            open_price = float(i['open'])
            high_price = float(i['high'])
            low_price = float(i['low'])

            if bound_unit_up == '$':
                if (high_price - open_price >= bound_up) and (open_price - low_price >= bound_up):
                    if data['min_interval'] == '60':
                        output = minute_shares_polygon(symbol=symbol, timeframe=timeframe, open_price=open_price,
                                                    date=i['time'], bound_up=bound_up, bound_low=bound_low)
                    elif data['min_interval'] == '1':
                        output = second_shares_polygon(symbol=symbol, timeframe=timeframe, open_price=open_price,
                                                    date=i['time'], bound_up=bound_up, bound_low=bound_low)
                elif high_price - open_price >= bound_up:
                    output = '1'
                elif open_price - low_price >= bound_low:
                    output = '0'
                else:
                    output = '2'
            elif bound_unit_up == '%':
                if (high_price - open_price >= (open_price / 100 * bound_up)) and (open_price - low_price >= (open_price / 100 * bound_low)):
                    if data['min_interval'] == '60':
                        output = minute_shares_polygon(symbol=symbol, timeframe=timeframe, open_price=open_price,
                                                    date=i['time'], bound_up=(open_price / 100 * bound_up), bound_low=(open_price / 100 * bound_low))
                    elif data['min_interval'] == '1':
                        output = second_shares_polygon(symbol=symbol, timeframe=timeframe, open_price=open_price,
                                                    date=i['time'], bound_up=(open_price / 100 * bound_up), bound_low=(open_price / 100 * bound_low))
                elif high_price - open_price >= (open_price / 100 * bound_up):
                    output = '1'
                elif open_price - low_price >= (open_price / 100 * bound_low):
                    output = '0'
                else:
                    output = '2'

            output_data.append({
                'time': i['time'],
                'output': output,
                'open': str(open_price),
                'close': str(i['close']),
                'high': str(high_price),
                'low': str(low_price),
                'trade': i['trade'],
                'volume': i['volume']
            })
            import time
            time.sleep(0.01)
            try:
                date_log = DateLog.objects.get(task_id=shares_polygon_async_task.request.id)
                date_log.delete()
            except:
                print('Ничего не найденно по такому ID')

        filtered_output_data = []
        if data['pre'] == 'in':
            if timeframe == '1 day':
                filtered_output_data = output_data
            else:
                for item in output_data:
                    time = datetime.strptime(item['time'], '%Y-%m-%d %H:%M:%S')
                    if (time.hour > 9 or (time.hour == 9 and time.minute >= 30)) and \
                            (time.hour < 15 or (time.hour == 15 and time.minute <= 30)):
                        filtered_output_data.append(item)
        elif data['pre'] == 'pre':
            filtered_output_data = output_data

        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Trade', 'Volume']
        for col_index, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_index, value=header)

        for row_index, item in enumerate(filtered_output_data, 2):
            row_data = [item['time'], item['output'], item['open'].replace(',', '.'), item['close'].replace(',', '.'), item['high'].replace(',', '.'), item['low'].replace(',', '.'), item['trade'],
                        item['volume']]
            for col_index, value in enumerate(row_data, 1):
                ws.cell(row=row_index, column=col_index, value=value)

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        file_path = f'{symbol}_{timeframe}_{bound_up}{bound_unit_up}_{bound_low}{bound_unit_low}_{start_date}_{end_date}(Polygon).xlsx'
        file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
        with open(file_path, 'wb') as file:
            file.write(output_buffer.read())

        if timeframe == '1 hour':
            input_file_path = file_path
            output_file_path = f'{symbol}_1_hour_{bound_up}{bound_unit_up}_{bound_low}{bound_unit_low}_{start_date}_{end_date}(Polygon).xlsx'

            data_processor = DataProcessor(input_file_path, output_file_path, bound_up=bound_up, bound_low=bound_low, symbol=symbol,
                                        bound_unit_up=bound_unit_up, bound_unit_low=bound_unit_low, min_interval=data['min_interval'])
            data_processor.load_data()
            data_processor.save_output_to_excel()
            task = Task.objects.get(user=data['us'], is_running=True)
            task.is_running = False
            task.save()
            return output_file_path, output_data
        else:
            task = Task.objects.get(user=data['us'], is_running=True)
            task.is_running = False
            task.save()
            return file_path, output_data


def send_notification_at_time(datetime_str):
    current_time = datetime.now()

    try:
        target_time = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
    except ValueError:
        print("Неверный формат даты и времени. Используйте формат 'ГГГГ-ММ-ДД ЧЧ:ММ'")
        return False

    if target_time <= current_time:
        print('Время пришло')
        return True

    time_until_notification = target_time - current_time
    delay = time_until_notification.total_seconds()

    if delay > 0:
        print(f"Уведомление будет отправлено через {time_until_notification}")
        time.sleep(delay)
        notification.notify(
            title='Уведомление',
            message='Пора!',
            app_name='ВашеПриложение'
        )
        return True
    else:
        print("Указанное время уже прошло.")
        return False


def trade(symbol, api_key, secret_key, leverage, amount_usdt, position):
    client = Client(api_key, secret_key, testnet=True)

    price = client.futures_mark_price(symbol=symbol)['markPrice']
    change_leverage = client.futures_change_leverage(symbol=symbol, leverage=leverage)

    col_vo_usdt = round((float(amount_usdt) * int(leverage)) / float(price), 3)
    print(col_vo_usdt, change_leverage)
    if position == '1':
        limit_order_long = client.futures_create_order(
            symbol=symbol,
            side='BUY',
            positionSide='LONG',
            type='LIMIT',
            quantity=col_vo_usdt,
            timeInForce='GTC',
            price=round(float(price), 0))
        print(limit_order_long)
    elif position == '0':
        limit_order_short = client.futures_create_order(
            symbol=symbol,
            side='SELL',
            positionSide='SHORT',
            type='LIMIT',
            quantity=col_vo_usdt,
            timeInForce='GTC',
            price=round(float(price), 0))
        print(limit_order_short)
    else:
        print('Получин не опознаный символ')


@shared_task
def async_parse_file_task(file_path, user_id, symbol, amount_usdt, leverage, api_key, secret_key):
    workbook = openpyxl.load_workbook(f'media/{file_path}')
    sheet = workbook['Sheet']
    data = []
    user = User.objects.get(pk=user_id)

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    current_time = datetime.strptime('00:00', '%H:%M')
    increments = {
        1441: 0.0166666667,
        481: 0.05,
        289: 0.0833333333,
        97: 0.25,
        49: 0.5,
        25: 1,
        13: 2,
        7: 4,
        5: 6,
        4: 8,
        3: 12,
        2: 24
    }

    for row in data:
        if len(row) in increments:
            date = row[0]
            current_increment = increments[len(row)]

            for i in row[1:]:
                combined_datetime = datetime.combine(date, current_time.time())
                position = i

                formatted_datetime = combined_datetime.strftime('%Y-%m-%d %H:%M')
                data_entry = DataEntry(user=user, date=formatted_datetime, position=position, symbol=symbol,
                                    amount_usdt=amount_usdt, leverage=leverage, api_key=api_key,
                                    secret_key=secret_key)
                data_entry.save()

                current_time += timedelta(hours=current_increment)
        else:
            break

    current_datetime = datetime.now()
    entries_to_process = DataEntry.objects.filter(user=user_id, is_completed=False)
    for entry in entries_to_process:
        print(entry.date, entry.position, entry.symbol, entry.amount_usdt, entry.leverage, entry.api_key,
              entry.secret_key)
        entry_datetime = datetime.strptime(entry.date, '%Y-%m-%d %H:%M')
        if entry_datetime > current_datetime:
            if send_notification_at_time(f"{entry_datetime.strftime('%Y-%m-%d %H:%M')}"):
                print('yes')
                trade(symbol=entry.symbol, api_key=entry.api_key, secret_key=entry.secret_key, leverage=entry.leverage,
                      amount_usdt=entry.amount_usdt, position=entry.position)
                entry.is_completed = True
                entry.save()


def minute_shares_yfinance(symbol, timeframe, open_price, date, bound_up, bound_low):
    interval_mapping = {
        '1m': 0.0166666667,
        '2m': 0.0333333333333333,
        '5m': 0.05,
        '15m': 0.0833333333,
        '30m': 0.25,
        '90m': 1.5,
        '1h': 1.0,
        '1d': 24.0,
        '5d': 120.0,
        '1wk': 168.0,
        '1mo': 720.0,
        '3mo': 2160.0
    }
    start_date = date
    if len(date) == 10:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
    end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
    data = yf.download(symbol, start=start_date_datetime, end=end_date_datetime, interval='1m')

    csv_filename = "historical_data_1m.csv"

    data.to_csv(csv_filename)

    data = pd.read_csv(csv_filename)
    data_dict = data.to_dict(orient="records")

    if os.path.exists(csv_filename):
        os.remove(csv_filename)
        print(f"Файл {csv_filename} удален.")

    for i in data_dict:
        if float(i['High']) - open_price >= bound_up:
            return '1'
        elif open_price - float(i['Low']) >= bound_low:
            return '0'


@shared_task
def shares_yfinance_async_task(data):
    symbol = data['symbol'].upper()
    start_date = data['start_data']
    end_date = data['end_data']
    end_date_inclusive = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    timeframe = data['interval']
    bound_up = float(data['bound_up'])
    bound_unit_up = data['bound_unit_up']
    bound_low = float(data['bound_up'])
    bound_unit_low = data['bound_unit_up']

    data = yf.download(symbol, start=start_date, end=end_date_inclusive, interval=timeframe)

    csv_filename = "historical_data.csv"

    data.to_csv(csv_filename)

    data = pd.read_csv(csv_filename)
    if timeframe == '1d' or timeframe == '5d' or timeframe == '1wk' or timeframe == '1mo' or timeframe == '3mo':
        name = 'Date'
    else:
        data["Datetime"] = data["Datetime"].str.replace("-04:00", "")
        name = 'Datetime'
    data_dict = data.to_dict(orient="records")

    if os.path.exists(csv_filename):
        os.remove(csv_filename)
        print(f"Файл {csv_filename} удален.")

    output_data = []
    for i in data_dict:
        if timeframe == '1d' or timeframe == '5d' or timeframe == '1wk' or timeframe == '1mo' or timeframe == '3mo':
            parsed_date = datetime.strptime(i[name], "%Y-%m-%d")
        else:
            parsed_date = datetime.strptime(i[name], "%Y-%m-%d %H:%M:%S")
        aware_date = timezone.make_aware(parsed_date)
        DateLog.objects.create(date=aware_date.date(), task_id=shares_yfinance_async_task.request.id)

        high_minus_open = float(i['High']) - float(i['Open'])
        open_minus_low = float(i['Open']) - float(i['Low'])

        if bound_unit_up == '%' and float(i['Open']) != 0:
            bound_up = (float(i['Open']) / 100) * bound_up
        
        if bound_unit_low == '%' and float(i['Open']) != 0:
            bound_low = (float(i['Open']) / 100) * bound_low
            
        time = i[name]

        if high_minus_open >= bound_up and open_minus_low >= bound_low:
            output = minute_shares_yfinance(
                symbol=symbol,
                timeframe=timeframe,
                open_price=float(i['Open']),
                date=time,
                bound_up=bound_up,
                bound_low=bound_low,
            )
        elif high_minus_open >= bound_up:
            output = '1'
        elif open_minus_low >= bound_low:
            output = '0'
        else:
            output = '2'

        ope = i['Open']
        close = i['Close']
        high = i['High']
        low = i['Low']
        volume = i['Volume']

        output_data.append({
            'time': time,
            'output': output,
            'open': ope,
            'close': close,
            'high': high,
            'low': low,
            'volume': volume
        })

        import time
        time.sleep(0.01)
        try:
            date_log = DateLog.objects.get(task_id=shares_yfinance_async_task.request.id)
            date_log.delete()
        except:
            print('Nothing found for this ID')

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low', 'Volume']
    for col_index, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_index, value=header)

    for item in output_data:
        row_data = [item['time'], item['output'], item['open'], item['close'], item['high'], item['low'],
                    item['volume']]
        ws.append(row_data)
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    file_path = f'{symbol}_{timeframe}_{bound_up}{bound_unit_up}_{bound_low}{bound_unit_low}_{start_date}_{end_date}(YahooFinance).xlsx'
    file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
    with open(file_path, 'wb') as file:
        file.write(output_buffer.read())
    task = Task.objects.get(is_running=True)
    task.is_running = False
    task.save()
    return file_path, output_data


def read_csv_data(file_path):
    data = []
    with open(file_path, 'r') as file:
        csv_reader = csv.reader(file)
        next(csv_reader)
        for row in csv_reader:
            timestamp = int(row[0])
            utc_datetime = datetime.utcfromtimestamp(timestamp)
            ny_timezone = pytz.timezone('America/New_York')
            ny_datetime = utc_datetime.replace(tzinfo=pytz.utc).astimezone(ny_timezone)
            formatted_date_time = ny_datetime.strftime('%Y-%m-%d %H:%M:%S')
            entry = {
                'date_time': formatted_date_time,
                'open': float(row[1]),
                'high': float(row[2]),
                'low': float(row[3]),
                'close': float(row[4])
            }
            data.append(entry)
    return data


def get_data_for_datetime_range(start_datetime, end_datetime, data):
    result = []
    for entry in data:
        entry_datetime = datetime.strptime(entry['date_time'], '%Y-%m-%d %H:%M:%S')
        if start_datetime <= entry_datetime <= end_datetime:
            result.append(entry)
    return result


def minute_TV(timeframe, open_price, bound, date, data):
    interval_mapping = {
        '1m': 0.0166666666666667,
        '3m': 0.05,
        '5m': 0.0833333333333333,
        '15m': 0.25,
        '30m': 0.5,
        '45m': 0.75,
        '1h': 1.0,
        '2h': 2.0,
        '3h': 3.0,
        '4h': 4.0,
        '1d': 24.0,
        '1wk': 168.0,
        '1mo': 720.0,
        '3mo': 2160.0,
        '6mo': 4320.0,
        '12mo': 8640.0
    }
    start_date_datetime = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    end_date_datetime = start_date_datetime + timedelta(hours=float(interval_mapping[timeframe]))

    data_for_desired_date_range = get_data_for_datetime_range(start_date_datetime, end_date_datetime, data)

    if data_for_desired_date_range:
        for i in data_for_desired_date_range:
            print(i)
            if float(i['high']) - open_price >= bound:
                return '1'
            elif open_price - float(i['low']) >= bound:
                return '0'
    else:
        print(f"Data for the date range from {start_date_datetime} to {end_date_datetime} not found.")


@shared_task
def tradingview_async_task(datas):
    symbol = datas['symbol']
    interval = datas['interval']
    bound = float(datas['bound'])
    bound_unit = datas['bound_unit']
    start_datetime = datas['start_date']
    end_datetime = datas['end_date']
    print(start_datetime)
    print(end_datetime)

    start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
    end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
    print(datas["file_for_big_bar"])
    data = read_csv_data(datas["file_for_big_bar"])
    data_for_desired_datetime_range = get_data_for_datetime_range(start_datetime, end_datetime, data)

    output_data = []

    if data_for_desired_datetime_range:
        for i in data_for_desired_datetime_range:
            parsed_date = datetime.strptime(i['date_time'], "%Y-%m-%d %H:%M:%S")
            aware_date = timezone.make_aware(parsed_date)
            DateLog.objects.create(date=aware_date.date(), task_id=tradingview_async_task.request.id)
            high_minus_open = float(i['high']) - float(i['open'])
            open_minus_low = float(i['open']) - float(i['low'])
            time = i['date_time']

            if bound_unit == '$':
                bound_value = bound
            elif bound_unit == '%':
                bound_value = (float(i['open']) / 100) * bound

            if high_minus_open >= bound_value and open_minus_low >= bound_value:
                data_for_1 = read_csv_data(datas["file_for_small_bar"])
                output = minute_TV(open_price=float(i['open']), bound=bound_value, date=i['date_time'],
                                   timeframe=interval, data=data_for_1)
                print(output)
            elif high_minus_open >= bound_value:
                output = '1'
            elif open_minus_low >= bound_value:
                output = '0'
            else:
                output = '2'

            ope = i['open']
            close = i['close']
            high = i['high']
            low = i['low']

            output_data.append({'time': time, 'output': output, 'open': ope, 'close': close, 'high': high, 'low': low})
            import time
            time.sleep(0.01)
            try:
                date_log = DateLog.objects.get(task_id=tradingview_async_task.request.id)
                date_log.delete()
            except:
                print('Ничего не найденно по такому ID')
    else:
        print(f"Data for the time range from {start_datetime} to {end_datetime} not found.")

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Date', 'Output', 'Open', 'Close', 'High', 'Low']
    for col_index, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_index, value=header)

    for item in output_data:
        row_data = [item['time'], item['output'], item['open'], item['close'], item['high'], item['low']]
        ws.append(row_data)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    file_path = f'{symbol}_{interval}_{bound}{bound_unit}_{start_datetime}_{end_datetime}(Tradingview).xlsx'
    file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
    with open(file_path, 'wb') as file:
        file.write(output_buffer.read())
    task = Task.objects.get(user=data['us'], is_running=True)
    task.is_running = False
    task.save()
    return file_path



def check_crossing_low(avg, previous_high, previous_low, date, symbol, timeframe):
    try:
        interval_mapping = {
            '1 minute': 0.0166666667,
            '5 minute': 0.0833333333,
            '15 minute': 0.25,
            '30 minute': 0.5,
            '45 minute': 0.75,
            '1 hour': 1.0,
            '2 hour': 2.0,
            '3 hour': 3.0,
            '4 hour': 4.0,
            '5 hour': 5.0,
            '6 hour': 6.0,
            '7 hour': 7.0,
            '8 hour': 8.0,
            '9 hour': 9.0,
            '10 hour': 10.0,
            '11 hour': 11.0,
            '12 hour': 12.0,
            '1 day': 24.0,
            '1 week': 168.0,
            '1 month': 720.0,
            '1 year': 8760
        }
        
        start_date = date
        print(start_date)
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000
        url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz'
        
        response = requests.get(url).json()['results']
        print(avg, previous_high, previous_low, start_unix_timestamp_milliseconds, end_unix_timestamp_milliseconds, symbol, timeframe)
        crossed_avg = False

        for i, candle in enumerate(response):
            print(candle)
            if crossed_avg == False and candle['o'] < avg and candle['h'] > avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['o'] > avg and candle['l'] < avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['l'] < previous_low:
                output = '2'
                status = 'NOT ACTIVE'
                return output, status, crossed_avg 
            elif crossed_avg and candle['l'] < previous_low:
                output = '0'
                status = 'ACTIVE'
                return output, status, crossed_avg 
    except Exception as e:
        print(e) 
    return '1/0', 'ACTIVE', crossed_avg


def check_crossing_low_or_high(avg, previous_high, previous_low, date, symbol, timeframe):
    try:
        interval_mapping = {
            '1 minute': 0.0166666667,
            '5 minute': 0.0833333333,
            '15 minute': 0.25,
            '30 minute': 0.5,
            '45 minute': 0.75,
            '1 hour': 1.0,
            '2 hour': 2.0,
            '3 hour': 3.0,
            '4 hour': 4.0,
            '5 hour': 5.0,
            '6 hour': 6.0,
            '7 hour': 7.0,
            '8 hour': 8.0,
            '9 hour': 9.0,
            '10 hour': 10.0,
            '11 hour': 11.0,
            '12 hour': 12.0,
            '1 day': 24.0,
            '1 week': 168.0,
            '1 month': 720.0,
            '1 year': 8760
        }
        
        start_date = date
        print(start_date)
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000
        url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz'
        
        response = requests.get(url).json()['results']
        print(avg, previous_high, previous_low, start_unix_timestamp_milliseconds, end_unix_timestamp_milliseconds, symbol, timeframe)
        crossed_avg = False

        for i, candle in enumerate(response):
            print(candle)
            if crossed_avg == False and candle['o'] < avg and candle['h'] > avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['o'] > avg and candle['l'] < avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['l'] < previous_low:
                output = '2'
                status = 'NOT ACTIVE'
                return output, status 
            elif crossed_avg == False and candle['h'] > previous_high:
                output = '2'
                status = 'NOT ACTIVE'
                return output, status 
            elif crossed_avg and candle['h'] > previous_high:
                output = '1'
                status = 'ACTIVE'
                return output, status 
            elif crossed_avg and candle['l'] < previous_low:
                output = '0'
                status = 'ACTIVE'
                return output, status 
    except Exception as e:
        print(e)   


def check_crossing_high(avg, previous_high, previous_low, date, symbol, timeframe):
    try:
        interval_mapping = {
            '1 minute': 0.0166666667,
            '5 minute': 0.0833333333,
            '15 minute': 0.25,
            '30 minute': 0.5,
            '45 minute': 0.75,
            '1 hour': 1.0,
            '2 hour': 2.0,
            '3 hour': 3.0,
            '4 hour': 4.0,
            '5 hour': 5.0,
            '6 hour': 6.0,
            '7 hour': 7.0,
            '8 hour': 8.0,
            '9 hour': 9.0,
            '10 hour': 10.0,
            '11 hour': 11.0,
            '12 hour': 12.0,
            '1 day': 24.0,
            '1 week': 168.0,
            '1 month': 720.0,
            '1 year': 8760
        }
        
        start_date = date
        print(start_date)
        start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        if start_date_datetime.time() == datetime.strptime("00:00:00", "%H:%M:%S").time():
            start_date_datetime = start_date_datetime.replace(hour=9, minute=30, second=0)
        else:
            start_date_datetime = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
        ny_timezone = pytz.timezone('America/New_York')
        start_date_datetime = ny_timezone.localize(start_date_datetime)
        end_date_datetime = start_date_datetime + timedelta(hours=interval_mapping[timeframe])
        start_unix_timestamp_milliseconds = int(start_date_datetime.timestamp()) * 1000
        end_unix_timestamp_milliseconds = int(end_date_datetime.timestamp()) * 1000
        url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/1/minute/{start_unix_timestamp_milliseconds}/{end_unix_timestamp_milliseconds}?adjusted=true&sort=asc&limit=50000&apiKey=EH2vpdYrp_dt3NHfcTjPhu0JOKKw0Lwz'
        
        response = requests.get(url).json()['results']
        print(avg, previous_high, previous_low, start_unix_timestamp_milliseconds, end_unix_timestamp_milliseconds, symbol, timeframe)
        crossed_avg = False

        for i, candle in enumerate(response):
            print(candle)
            if crossed_avg == False and candle['o'] > avg and candle['l'] < avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['o'] > avg and candle['l'] < avg:
                crossed_avg = True
                print('Было пересечение средины')
            elif crossed_avg == False and candle['h'] > previous_high:
                output = '2'
                status = 'NOT ACTIVE'
                return output, status, crossed_avg
            elif crossed_avg and candle['h'] > previous_high:
                output = '1'
                status = 'ACTIVE'
                crossed_avg = False
                return output, status, crossed_avg
            
    except Exception as e:
        print(e)
    return '1/0', 'ACTIVE', crossed_avg
            


@shared_task
def shares_polygon_new_async_task(data):
    symbol = data['symbol']
    timeframe = data['timeframe']
    interval_start = float(data['interval_start'])
    interval_end = float(data['interval_end'])
    start_date = data['start_date']
    end_date = data['end_date']
    api_key = data['api_key']
    
    if timeframe == '4 hour' or timeframe == '1 hour':
        formating = FormatingDataServiceNew(symbol=symbol, timeframe=timeframe, interval_start=interval_start, interval_end=interval_end,  start_date=start_date, end_date=end_date, api_key=api_key)
        file_path, output_data = formating.save_output_to_excel()
        # task = Task.objects.get(user=data['us'], is_running=True)
        # task.is_running = False
        # task.save()
        return file_path, output_data
    else:
        interval_parts = timeframe.split()
        url = f'https://api.polygon.io/v2/aggs/ticker/{symbol}/range/{interval_parts[0]}/{interval_parts[1]}/{start_date}/{end_date}?adjusted=true&sort=asc&limit=50000&apiKey={api_key}'

        response = requests.get(url).json()['results']

        avg = 0.0
        previous_high = 0.0
        previous_low = 0.0

        output_data = []
        for i in range(len(response) - 1):
            previous_candle = response[i - 1]
            candle = response[i]
            next_candle = response[i + 1]
            
            unix_timestamp_seconds = candle['t'] / 1000
            unix_datetime = datetime.fromtimestamp(unix_timestamp_seconds, pytz.utc)
            ny_timezone = pytz.timezone('America/New_York')
            ny_datetime = unix_datetime.astimezone(ny_timezone)
            time = ny_datetime.strftime("%Y-%m-%d %H:%M:%S")
            open_price = candle['o']
            high = candle['h']
            low = candle['l']
            close = candle['c'] 
            volume = candle['v']
            trade = candle['n']
            next_open = next_candle['o']
            amplitude = ((high - low) / low) * 100
            print(candle)
            unix_timestamp_seconds = next_candle['t'] / 1000
            unix_datetime = datetime.fromtimestamp(unix_timestamp_seconds, pytz.utc)
            ny_timezone = pytz.timezone('America/New_York')
            ny_datetime_next = unix_datetime.astimezone(ny_timezone)
            next_time = ny_datetime_next.strftime("%Y-%m-%d %H:%M:%S")
            if interval_start <= amplitude <= interval_end:
                avg = (high + low) / 2
                next_high = next_candle['h']
                next_low = next_candle['l']
                if high < next_high and next_low > avg:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    status = 'NOT ACTIVE'
                    output = '2'
                    print(status, output)
                elif next_open > avg and next_low < low and next_high < high:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    status = 'ACTIVE'
                    output = '0'
                    print(status, output)
                elif next_open < avg and next_high > high and next_low > low:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    status = 'ACTIVE'
                    output = '1'
                    print(status, output)
                elif next_high > high and next_low < low:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    output, status  = check_crossing_low_or_high(avg, high, low, next_time, symbol, timeframe)
                    print(status, output)
                elif next_high > avg and next_low < avg and next_high < high and next_low > low:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    previous_high = high
                    previous_low = low
                    for j in response[i:-1]:
                        if j['h'] > previous_high:
                            status = 'ACTIVE'
                            output = '1'
                            print(status, output)
                            break
                        elif j['l'] < previous_low:
                            status = 'ACTIVE'
                            output = '0'
                            print(status, output)
                            break
                elif low > next_low and next_high < avg:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    status = 'NOT ACTIVE'
                    output = '2'
                    print(status, output)
                elif high < next_high and next_low < avg:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    output, status, crossed_avg  = check_crossing_high(avg, high, low, next_time, symbol, timeframe)
                    print(status, output)
                    if output == '1/0' and status == 'ACTIVE' and crossed_avg == True:
                        previous_high = high
                        previous_low = low
                        for j in response[i:-1]:
                            if j['h'] > previous_high:
                                status = 'ACTIVE'
                                output = '1'
                                print(status, output)
                                break
                            elif j['l'] < previous_low:
                                status = 'ACTIVE'
                                output = '0'
                                print(status, output)
                                break
                elif low > next_low and next_high > avg:
                    print(f'high - {high}\next_high - {next_high}\next_low - {next_low}\navg - {avg}\nlow - {low}')
                    output, status, crossed_avg = check_crossing_low(avg, high, low, next_time, symbol, timeframe)
                    print(status, output)
                    if output == '1/0' and status == 'ACTIVE' and crossed_avg == True:
                        previous_high = high
                        previous_low = low
                        for j in response[i:-1]:
                            if j['h'] > previous_high:
                                status = 'ACTIVE'
                                output = '1'
                                print(status, output)
                                break
                            elif j['l'] < previous_low:
                                status = 'ACTIVE'
                                output = '0'
                                print(status, output)
                                break
                else:
                    status = 'NOT ACTIVE'
                    output = '2'
                    print(status, output)
            else:
                status = 'NOT ACTIVE'
                output = '2'
                print(status, output)
                
            output_data.append({
                'time': ny_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                'status': status,
                'output': output,
                'open': candle['o'],
                'close': candle['c'],
                'high': candle['h'],
                'low': candle['l'],
                'trade': candle['n'],
                'volume': candle['v']
            })
            
        print(output_data)    
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Date', 'Status', 'Output', 'Open', 'Close', 'High', 'Low', 'Trade', 'Volume']
        for col_index, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_index, value=header)

        for item in output_data:
            row_data = [item['time'], item['status'], item['output'], item['open'], item['close'], item['high'], item['low'], item['trade'], item['volume']]
            ws.append(row_data)

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        file_path = f'{symbol}_{timeframe}_{interval_start}%_{interval_end}%{start_date}_{end_date}(Polugon).xlsx'
        file_path = file_path.replace(':', '_').replace('?', '_').replace(' ', '_')
        with open(file_path, 'wb') as file:
            file.write(output_buffer.read())
        
        return file_path, output_data
